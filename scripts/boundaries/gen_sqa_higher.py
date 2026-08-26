"""SQA Highers, modelled as components — the same treatment as the Advanced
Highers in gen_sqa_ah.py, and for the same reasons.

SQA publishes grade boundaries for the whole course only, so each component's
boundary is its share of the course cut-off and every row is written with
derived_from_course = TRUE. The component names and max marks are SQA's own,
from the Assessment and Component Marks tables, and each course is checked to
sum to the course maximum against a second, separate publication.

Graded A-D: no A*, no E, No Award below D.

2024 and 2025 only. In 2022 and 2023 the courses ran in a modified form —
Higher Geography was 70 marks rather than 110, the sciences 120 rather than 150
— which is a different set of components.

One wrinkle worth naming: SQA is not consistent about its own component names
across years. Higher German's coursework is "Assignment: Writing" in 2024 and
"Assignment - Writing" in 2025. The names are normalised before comparison, so
a punctuation change does not read as a course restructure.
"""

import sys, os  # noqa: E402
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _paths import DOCS, MIGRATIONS, REPO, require_docs  # noqa: E402
import os
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

TELOS = REPO
SP = DOCS
SHEET = "Higher"

# Written papers first, then the oral, then the coursework.
ORDER = {"exam": 0, "oral": 1, "coursework": 2}

YEARS = ["2022", "2023", "2024", "2025"]


def norm_name(name):
    """SQA's own component names drift in punctuation between years."""
    # Commas go too: 2023 writes "Reading for Understanding  Analysis and
    # Evaluation" where every other year writes it with a comma.
    return re.sub(r"[\s:_,-]+", " ", str(name)).strip().lower()


# normalised SQA name -> (stored code, label a student reads, how it is marked)
COMPONENTS = {
    "paper 1 (multiple choice)":   ("Paper 1", "Paper 1 (Multiple Choice)", "exam"),
    "paper 2":                     ("Paper 2", "Paper 2", "exam"),
    "paper 1 (non calculator)":    ("Paper 1", "Paper 1 (Non-calculator)", "exam"),
    "paper 2 (calculator)":        ("Paper 2", "Paper 2 (Calculator)", "exam"),
    "question paper":              ("Question Paper", "Question Paper", "exam"),
    "assignment":                  ("Assignment", "Assignment", "coursework"),
    "assignment writing":          ("Assignment", "Assignment: Writing", "coursework"),
    "reading":                     ("Reading", "Reading", "exam"),
    "listening":                   ("Listening", "Listening", "exam"),
    "directed writing":            ("Directed Writing", "Directed Writing", "exam"),
    "performance talking":         ("Talking", "Performance: Talking", "oral"),
    "physical and human environments":
        ("Paper 1", "Physical and Human Environments", "exam"),
    "global issues and geographical skills":
        ("Paper 2", "Global Issues and Geographical Skills", "exam"),
    "reading for understanding analysis and evaluation":
        ("Paper 1", "Reading for Understanding, Analysis and Evaluation", "exam"),
    "critical reading":            ("Paper 2", "Critical Reading", "exam"),
    "portfolio writing":           ("Portfolio", "Portfolio: Writing", "coursework"),
}

SUBJECTS = {
    "Biology":     ("Biology (H)", "Biology", "#5E9E6B"),
    "Chemistry":   ("Chemistry (H)", "Chemistry", "#5E8B7E"),
    "Physics":     ("Physics (H)", "Physics", "#5E8B7E"),
    "Mathematics": ("Maths (H)", "Maths", "#C9A227"),
    "Economics":   ("Economics (H)", "Economics", "#C08A3E"),
    "Geography":   ("Geography (H)", "Geography", "#6E8F5E"),
    "French":      ("French (H)", "French", "#4C7EF3"),
    "German":      ("German (H)", "German", "#C08A3E"),
    "Spanish":     ("Spanish (H)", "Spanish", "#D06A5A"),
    # The most-taken Higher in Scotland, and the first subject in the catalogue
    # with no sibling at any other level.
    "English":     ("English (H)", "English", "#8A6FA8"),
}

SCI_TOPICS = {
    "Biology (H)": ["DNA and the Genome", "Metabolism and Survival",
                    "Sustainability and Interdependence"],
    "Chemistry (H)": ["Chemical Changes and Structure", "Nature's Chemistry",
                      "Chemistry in Society", "Researching Chemistry"],
    "Physics (H)": ["Our Dynamic Universe", "Particles and Waves", "Electricity"],
}
TOPICS = {}
for _k, _t in SCI_TOPICS.items():
    TOPICS[_k] = {"Paper 1": _t, "Paper 2": _t, "Assignment": ["Assignment"]}
TOPICS["Maths (H)"] = {
    "Paper 1": ["Algebraic and Trigonometric Skills", "Geometric Skills",
                "Calculus Skills", "Algebraic and Geometric Skills"],
    "Paper 2": ["Algebraic and Trigonometric Skills", "Geometric Skills",
                "Calculus Skills", "Reasoning Skills"],
}
TOPICS["Economics (H)"] = {
    "Question Paper": ["Economics of the Market", "UK Economic Activity",
                       "Global Economic Activity"],
    "Assignment": ["Economics Assignment"],
}
TOPICS["Geography (H)"] = {
    "Paper 1": ["Atmosphere", "Hydrosphere", "Lithosphere", "Biosphere",
                "Population", "Rural Land Use", "Urban"],
    "Paper 2": ["River Basin Management", "Development and Health",
                "Global Climate Change", "Energy", "Geographical Skills"],
    "Assignment": ["Geographical Assignment"],
}
TOPICS["English (H)"] = {
    "Paper 1": ["Reading for Understanding", "Analysis", "Evaluation",
                "Summarising", "Comparison of Passages"],
    "Paper 2": ["Scottish Text", "Critical Essay: Drama",
                "Critical Essay: Prose", "Critical Essay: Poetry",
                "Critical Essay: Film and Television Drama",
                "Critical Essay: Language"],
    "Portfolio": ["Broadly Creative Writing", "Broadly Discursive Writing"],
}
for _k in ("French (H)", "German (H)", "Spanish (H)"):
    TOPICS[_k] = {
        "Reading": ["Reading Comprehension"],
        "Listening": ["Listening Comprehension"],
        "Directed Writing": ["Directed Writing"],
        "Talking": ["Presentation", "Conversation"],
        "Assignment": ["Written Assignment"],
    }


def component_table(year):
    """{subject: (course_max, [(code, label, kind, max), ...])} for one year.

    Four layouts across four files. 2022 leads with a Qualification Number
    column and puts its header on row 1; 2023 has no such column and heads on
    row 2; 2024 and 2025 head on row 3 and add an "Assessment Maximum Mark".
    So the header row is found by looking for the Subject cell rather than
    being assumed, and the subject column is wherever that cell turned up.

    Where the file carries no assessment total, the components are summed and
    that sum is checked against the course maximum in the grade boundary
    release — which is a stronger check than reading a total from the same
    file it is meant to validate.
    """
    wb = load_workbook(os.path.join(SP, "sqacomp_%s.xlsx" % year), data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i, subj_i = next(
        (i, [str(c) for c in r].index("Subject"))
        for i, r in enumerate(rows)
        if r and "Subject" in [str(c) for c in r])
    hdr = [("" if h is None else str(h)) for h in rows[hdr_i]]
    total_i = hdr.index("Assessment Maximum Mark") if "Assessment Maximum Mark" in hdr else None

    out = {}
    for r in rows[hdr_i + 1:]:
        if subj_i >= len(r) or r[subj_i] not in SUBJECTS:
            continue
        comps = []
        for i, h in enumerate(hdr):
            if not h.endswith(" Name") or i + 1 >= len(r):
                continue
            name, mx = r[i], r[i + 1]
            if not name or str(name).startswith("[") or not str(mx).isdigit():
                continue
            key = norm_name(name)
            if key not in COMPONENTS:
                raise SystemExit("unmapped component %r in %s %s"
                                 % (name, r[subj_i], year))
            code, label, kind = COMPONENTS[key]
            comps.append((code, label, kind, int(mx)))
        comps.sort(key=lambda c: (ORDER[c[2]], c[0]))
        total = int(r[total_i]) if total_i is not None and str(r[total_i]).isdigit()             else sum(m for _c, _l, _k, m in comps)
        out[r[subj_i]] = (total, comps)
    return out


def course_boundaries(year):
    wb = load_workbook(os.path.join(SP, "sqa_2025.xlsx"), data_only=True)
    ws = wb["Higher"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [("" if h is None else str(h)) for h in rows[3]]
    keys = ("Maximum Mark", "A Boundary", "B Boundary", "C Boundary", "D Boundary")
    cols = {k: hdr.index("%s %s" % (k, year)) for k in keys}
    out = {}
    for r in rows[4:]:
        if r[0] not in SUBJECTS:
            continue
        vals = [r[cols[k]] for k in keys]
        if all(str(v).isdigit() for v in vals):
            out[r[0]] = tuple(int(v) for v in vals)
    return out


def fmt(items, indent=" " * 26):
    out, line = [], ""
    for it in items:
        piece = '"%s", ' % it
        if len(line) + len(piece) > 74:
            out.append(indent + line.rstrip())
            line = ""
        line += piece
    if line:
        out.append(indent + line.rstrip().rstrip(","))
    return "\n".join(out)


# The 2025 shape is the catalogue's shape: one set of papers per qualification.
# Earlier years are read against it rather than replacing it.
CANON_YEAR = "2025"
canon_comps, canon_max = {}, {}
for sqa_name, (key, _l, _c) in SUBJECTS.items():
    _cmax, _comps = component_table(CANON_YEAR)[sqa_name]
    canon_comps[key] = _comps
    canon_max[key] = {code: mx for code, _l2, _k2, mx in _comps}

values, seeds, shape, skipped = [], [], canon_comps, []
for year in YEARS:
    comps_by_subj = component_table(year)
    bounds = course_boundaries(year)
    for sqa_name, (key, _label, _colour) in sorted(SUBJECTS.items()):
        if sqa_name not in comps_by_subj or sqa_name not in bounds:
            raise SystemExit("%s missing for %s" % (sqa_name, year))
        course_max, comps = comps_by_subj[sqa_name]
        bmax, ba, bb, bc, bd = bounds[sqa_name]

        # The two publications must agree about the course before anything is
        # derived from them. They are separate releases and could drift.
        assert bmax == course_max, "%s %s: boundary max %d vs component max %d" % (
            sqa_name, year, bmax, course_max)
        summed = sum(m for _c, _l, _k, m in comps)
        assert summed == course_max, "%s %s: components sum to %d, course is %d" % (
            sqa_name, year, summed, course_max)

        for code, _l2, _k2, cmax in comps:
            # A component only earns a row for a year in which it was the same
            # paper. In 2022 and 2023 these courses ran in a modified form —
            # coursework withdrawn, question papers resized — so a component
            # that changed size, or did not exist, is skipped for that year
            # rather than stored against a paper of a different length. The
            # engine then falls back to the median of the years it does have.
            if canon_max[key].get(code) != cmax:
                skipped.append((key, year, code, cmax, canon_max[key].get(code)))
                continue
            # Each year is pro-rated against ITS OWN course total, which is what
            # makes a modified year usable at all: in 2023 Higher Biology Paper 2
            # was 95 marks of a 120-mark course, not 95 of 150.
            share = cmax / course_max
            a, b, c, d = (round(x * share) for x in (ba, bb, bc, bd))
            assert a > b > c > d > 0, "%s %s %s: %s not descending" % (
                sqa_name, year, code, (a, b, c, d))
            assert a < cmax, "%s %s %s: A %d is not below max %d" % (
                sqa_name, year, code, a, cmax)
            values.append(
                "    ('%s', 'SQA', '%s', '%s', 'June', NULL, %d, %d, %d, %d, NULL, TRUE)"
                % (key, code, year, a, b, c, d))
            seeds.append('    ("%s", "SQA", "%s", "%s", "June", None, %d, %d, %d),'
                         % (key, code, year, a, b, c))
    print("%s: %d subjects" % (year, len(SUBJECTS)))

if skipped:
    print("skipped (component absent or a different size that year):")
    for key, year, code, got, want in skipped:
        print("   %-16s %s %-16s max %s, catalogue says %s" % (key, year, code, got, want))

entries = []
for sqa_name, (key, label, colour) in sorted(SUBJECTS.items(), key=lambda kv: kv[1][0]):
    comps = shape[key]
    paper_lines = "\n".join(
        '                {"code": "%s", "name": "%s", "max_marks": %d%s},'
        % (code, lab, mx, "" if kind == "exam" else ', "assessment": "%s"' % kind)
        for code, lab, kind, mx in comps)
    topic_lines = "\n".join(
        '                "%s": [\n%s],' % (code, fmt(TOPICS[key][code]))
        for code, _l, _k, _m in comps)
    entry = '        "%s": {\n' % key
    entry += '            "name": "%s",\n' % label
    entry += '            "color": "%s",\n' % colour
    entry += '            "level": "Higher",\n'
    entry += ('            # Graded A-D: no A* and no E. Component max marks are\n'
              '            # SQA\'s own; the boundaries are derived from the course\n'
              '            # cut-off, because SQA publishes none per component.\n')
    entry += '            "papers": [\n%s\n            ],\n' % paper_lines
    entry += ('            # 2022 and 2023 ran in a modified form with a different\n'
              '            # set of components.\n')
    entry += '            "years": ["SPEC", "2022", "2023", "2024", "2025"],\n'
    entry += '            "topics": {\n%s\n            },\n' % topic_lines
    entry += '        },\n'
    entries.append(entry)

sql = (
    "-- 039_sqa_higher_boundaries.sql\n"
    "-- SQA Highers: Biology, Chemistry, Physics, Mathematics, Economics,\n"
    "-- Geography, French, German, Spanish.\n"
    "--\n"
    "-- Same treatment as the Advanced Highers in migration 038, for the same\n"
    "-- reason: SQA publishes cut-off scores for the whole course and never per\n"
    "-- component, so each component's boundary is its share of the course\n"
    "-- cut-off and derived_from_course is TRUE on every row.\n"
    "--\n"
    "--     component_boundary = round(course_boundary * component_max / course_max)\n"
    "--\n"
    "-- The component max marks are NOT derived. They are SQA's own, from the\n"
    "-- Assessment and Component Marks tables, and every course is checked to\n"
    "-- sum to the course maximum published in the separate grade boundaries\n"
    "-- release before anything is computed from it.\n"
    "--\n"
    "-- Graded A-D: a_star and e_boundary are both NULL. Below D is No Award.\n"
    "--\n"
    "-- 2024 and 2025 only. In 2022 and 2023 these courses ran in a modified\n"
    "-- form - Geography 70 marks rather than 110, the sciences 120 rather than\n"
    "-- 150 - which is a different set of components, not the same course with\n"
    "-- different numbers.\n"
    "--\n"
    "-- Idempotent.\n\n"
    "DELETE FROM grade_boundaries WHERE board = 'SQA' AND subject LIKE '%% (H)';\n\n"
    "INSERT INTO grade_boundaries\n"
    "    (subject, board, paper_code, year, series,\n"
    "     a_star, a_boundary, b_boundary, c_boundary, d_boundary, e_boundary,\n"
    "     derived_from_course)\n"
    "VALUES\n%s;\n"
) % (",\n".join(values))

with open(os.path.join(TELOS, "migrations", "039_sqa_higher_boundaries.sql"),
          "w", encoding="utf-8", newline="\n") as fh:
    fh.write(sql)
with open(os.path.join(SP, "sqa_h_seed.txt"), "w", encoding="utf-8") as fh:
    fh.write("\n".join(seeds))
with open(os.path.join(SP, "sqa_h_entries.txt"), "w", encoding="utf-8") as fh:
    fh.write("".join(entries))
print("migration 039 written with %d rows" % len(values))
