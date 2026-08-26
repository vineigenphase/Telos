"""SQA Advanced Highers, modelled as components.

SQA is not like the English boards, and this is the one place in the catalogue
where Telos computes something the awarding body does not publish.

  * SQA publishes grade boundaries at COURSE level only: one maximum mark and
    A/B/C/D cut-off scores for the whole course. There is no published
    component boundary anywhere, for any subject, in any year.
  * SQA does publish, separately, the maximum mark of every component of every
    course (the "Assessment and Component Marks" tables). That is where the
    component structure here comes from — it is read, not inferred, and each
    course's components are checked to sum to the course maximum.

So the components are real and their max marks are real; the per-component
boundaries are DERIVED, by taking each component's share of the course cut-off:

    component_boundary = round(course_boundary * component_max / course_max)

Every row is written with derived_from_course = TRUE so this can never be
mistaken for published data. The approximation assumes a component is as hard
as the course as a whole, which is not exactly true — projects generally score
higher than question papers, so this reads slightly harsh on the project and
slightly generous on the paper. It is an estimate, labelled as one.

2022 to 2025. Several of these courses ran in a modified form in 2022 and 2023
— coursework withdrawn, question papers resized — so a component earns a row
for a year only when it was the same paper that year: same code, same max mark.
Advanced Higher Physics' question paper was 155 marks then against 120 now, so
it gets no row for those years rather than a boundary computed for a paper of a
different length. Every skip is printed on each run.

Subjects that were not modified — English among them — carry all four years.

An Advanced Higher is graded A-D with no E and no A*. a_star and e_boundary are
both NULL, and prediction.py reads those absences to build a ladder that runs
from D up to A.
"""

import sys, os  # noqa: E402
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _paths import DOCS, MIGRATIONS, REPO, require_docs  # noqa: E402
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

TELOS = REPO
SP = DOCS
SHEET = "Advanced_Higher"

# Written papers first, then the oral, then the coursework.
ORDER = {"exam": 0, "oral": 1, "coursework": 2}

import re


def norm_name(name):
    """SQA's component names drift in punctuation between years."""
    # Commas go too: 2023 writes "Reading for Understanding  Analysis and
    # Evaluation" where every other year writes it with a comma.
    return re.sub(r"[\s:_,-]+", " ", str(name)).strip().lower()


YEARS = ["2022", "2023", "2024", "2025"]

# SQA's component name -> the code Telos stores, the label a student reads, and
# how it is marked. "coursework" and "oral" get single-mark entry.
_RAW_COMPONENTS = {
    "Question Paper":              ("Question Paper", "Question Paper", "exam"),
    "Section 1 - Objective Test":  ("Section 1", "Section 1: Objective Test", "exam"),
    "Section 2":                   ("Section 2", "Section 2", "exam"),
    "Paper 1 (Non Calculator)":    ("Paper 1", "Paper 1 (Non-calculator)", "exam"),
    "Paper 2 (Calculator)":        ("Paper 2", "Paper 2 (Calculator)", "exam"),
    "Project":                     ("Project", "Project", "coursework"),
    "Project Folio: Section A":    ("Folio A", "Project-folio: Geographical Study", "coursework"),
    "Project Folio: Section B":    ("Folio B", "Project-folio: Geographical Issue", "coursework"),
    "Portfolio":                   ("Portfolio", "Portfolio", "coursework"),
    "Performance: Talking":        ("Talking", "Performance: Talking", "oral"),
    "Reading and Translation":     ("Reading", "Reading and Translation", "exam"),
    "Listening and Discursive Writing": ("Listening", "Listening and Discursive Writing", "exam"),
    # Advanced Higher English. SQA lists the question paper's two sections as
    # separate components, which is how a student meets them.
    "Literary Study":              ("Literary Study", "Literary Study", "exam"),
    "Textual Analysis":            ("Textual Analysis", "Textual Analysis", "exam"),
    "Portfolio: Writing":          ("Portfolio", "Portfolio: Writing", "coursework"),
    "Project: Dissertation":       ("Dissertation", "Project: Dissertation", "coursework"),
}

COMPONENTS = {norm_name(k): v for k, v in _RAW_COMPONENTS.items()}

SUBJECTS = {
    "Biology":     ("Biology (AH)", "Biology", "#5E9E6B"),
    "Chemistry":   ("Chemistry (AH)", "Chemistry", "#5E8B7E"),
    "Physics":     ("Physics (AH)", "Physics", "#5E8B7E"),
    "Mathematics": ("Maths (AH)", "Maths", "#C9A227"),
    "Economics":   ("Economics (AH)", "Economics", "#C08A3E"),
    "Geography":   ("Geography (AH)", "Geography", "#6E8F5E"),
    "French":      ("French (AH)", "French", "#4C7EF3"),
    "German":      ("German (AH)", "German", "#C08A3E"),
    "Spanish":     ("Spanish (AH)", "Spanish", "#D06A5A"),
    # Telos has English only at SQA, and now at both levels.
    "English":     ("English (AH)", "English", "#8A6FA8"),
}

TOPICS = {
    "Biology (AH)": {
        "Section 1": ["Cells and Proteins", "Organisms and Evolution",
                      "Investigative Biology"],
        "Section 2": ["Cells and Proteins", "Organisms and Evolution",
                      "Investigative Biology"],
        "Project": ["Research Project"],
    },
    "Chemistry (AH)": {
        "Section 1": ["Inorganic and Physical Chemistry", "Organic Chemistry and Instrumental Analysis",
                      "Researching Chemistry"],
        "Section 2": ["Inorganic and Physical Chemistry", "Organic Chemistry and Instrumental Analysis",
                      "Researching Chemistry"],
        "Project": ["Research Project"],
    },
    "Physics (AH)": {
        "Question Paper": ["Rotational Motion and Astrophysics", "Quanta and Waves",
                           "Electromagnetism", "Units, Prefixes and Uncertainties"],
        "Project": ["Research Project"],
    },
    "Maths (AH)": {
        "Paper 1": ["Methods in Algebra and Calculus", "Applications of Algebra and Calculus",
                    "Geometry, Proof and Systems of Equations"],
        "Paper 2": ["Methods in Algebra and Calculus", "Applications of Algebra and Calculus",
                    "Geometry, Proof and Systems of Equations"],
    },
    "Economics (AH)": {
        "Question Paper": ["Economics of the Market", "Global Economic Activity",
                           "The UK Economy", "Economic Data and Analysis"],
        "Project": ["Economics Project"],
    },
    "Geography (AH)": {
        "Question Paper": ["Geographical Methods and Techniques"],
        "Folio A": ["Geographical Study"],
        "Folio B": ["Geographical Issue"],
    },
}
TOPICS["English (AH)"] = {
    "Literary Study": ["Critical Essay on a Chosen Text", "Poetry", "Prose Fiction",
                       "Prose Non-fiction", "Drama", "Film and Television Drama",
                       "Language Study"],
    "Textual Analysis": ["Unseen Poetry", "Unseen Prose", "Unseen Drama",
                         "Comparative Analysis"],
    "Portfolio": ["Broadly Creative Writing", "Broadly Discursive Writing"],
    "Dissertation": ["Independent Literary Study"],
}
for _k in ("French (AH)", "German (AH)", "Spanish (AH)"):
    TOPICS[_k] = {
        "Reading": ["Reading Comprehension", "Translation into English"],
        "Listening": ["Listening Comprehension", "Discursive Writing"],
        "Talking": ["Presentation", "Discussion"],
        "Portfolio": ["Written Portfolio"],
    }


def component_table(year):
    """{subject: (course_max, [(code, label, kind, max), ...])} for one year.

    Four layouts across four files. 2022 leads with a Qualification Number
    column and puts its header on row 1; 2023 has no such column and heads on
    row 2; 2024 and 2025 head on row 3 and add an "Assessment Maximum Mark".
    So the header row is found by looking for the Subject cell rather than
    being assumed, and the subject column is wherever that cell turned up.

    Where the file carries no assessment total, the components are summed and
    that sum is checked against the course maximum in the grade boundary
    release — which is a stronger check than reading a total from the same
    file it is meant to validate.
    """
    wb = load_workbook(os.path.join(SP, "sqacomp_%s.xlsx" % year), data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i, subj_i = next(
        (i, [str(c) for c in r].index("Subject"))
        for i, r in enumerate(rows)
        if r and "Subject" in [str(c) for c in r])
    hdr = [("" if h is None else str(h)) for h in rows[hdr_i]]
    total_i = hdr.index("Assessment Maximum Mark") if "Assessment Maximum Mark" in hdr else None

    out = {}
    for r in rows[hdr_i + 1:]:
        if subj_i >= len(r) or r[subj_i] not in SUBJECTS:
            continue
        comps = []
        for i, h in enumerate(hdr):
            if not h.endswith(" Name") or i + 1 >= len(r):
                continue
            name, mx = r[i], r[i + 1]
            if not name or str(name).startswith("[") or not str(mx).isdigit():
                continue
            key = norm_name(name)
            if key not in COMPONENTS:
                raise SystemExit("unmapped component %r in %s %s"
                                 % (name, r[subj_i], year))
            code, label, kind = COMPONENTS[key]
            comps.append((code, label, kind, int(mx)))
        comps.sort(key=lambda c: (ORDER[c[2]], c[0]))
        total = int(r[total_i]) if total_i is not None and str(r[total_i]).isdigit()             else sum(m for _c, _l, _k, m in comps)
        out[r[subj_i]] = (total, comps)
    return out


def course_boundaries(year):
    """{subject: (max, a, b, c, d)} from the course-level grade boundary sheet."""
    wb = load_workbook(os.path.join(SP, "sqa_2025.xlsx"), data_only=True)
    ws = wb["Advanced_Higher"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [("" if h is None else str(h)) for h in rows[3]]
    cols = {k: hdr.index("%s %s" % (k, year)) for k in
            ("Maximum Mark", "A Boundary", "B Boundary", "C Boundary", "D Boundary")}
    out = {}
    for r in rows[4:]:
        if r[0] not in SUBJECTS:
            continue
        vals = [r[cols[k]] for k in
                ("Maximum Mark", "A Boundary", "B Boundary", "C Boundary", "D Boundary")]
        if not all(str(v).isdigit() for v in vals):
            continue
        out[r[0]] = tuple(int(v) for v in vals)
    return out


def fmt(items, indent=" " * 26):
    out, line = [], ""
    for it in items:
        piece = '"%s", ' % it
        if len(line) + len(piece) > 74:
            out.append(indent + line.rstrip())
            line = ""
        line += piece
    if line:
        out.append(indent + line.rstrip().rstrip(","))
    return "\n".join(out)


# The 2025 shape is the catalogue's shape: one set of papers per qualification.
# Earlier years are read against it rather than replacing it.
CANON_YEAR = "2025"
canon_comps, canon_max = {}, {}
for sqa_name, (key, _l, _c) in SUBJECTS.items():
    _cmax, _comps = component_table(CANON_YEAR)[sqa_name]
    canon_comps[key] = _comps
    canon_max[key] = {code: mx for code, _l2, _k2, mx in _comps}

values, seeds, shape, skipped = [], [], canon_comps, []
for year in YEARS:
    comps_by_subj = component_table(year)
    bounds = course_boundaries(year)
    for sqa_name, (key, _label, _colour) in sorted(SUBJECTS.items()):
        if sqa_name not in comps_by_subj or sqa_name not in bounds:
            raise SystemExit("%s missing for %s" % (sqa_name, year))
        course_max, comps = comps_by_subj[sqa_name]
        bmax, ba, bb, bc, bd = bounds[sqa_name]

        # The two publications must agree about the course before anything is
        # derived from them. They are separate releases and could drift.
        assert bmax == course_max, "%s %s: boundary max %d vs component max %d" % (
            sqa_name, year, bmax, course_max)
        summed = sum(m for _c, _l, _k, m in comps)
        assert summed == course_max, "%s %s: components sum to %d, course is %d" % (
            sqa_name, year, summed, course_max)

        for code, _l2, _k2, cmax in comps:
            # A component only earns a row for a year in which it was the same
            # paper. In 2022 and 2023 these courses ran in a modified form —
            # coursework withdrawn, question papers resized — so a component
            # that changed size, or did not exist, is skipped for that year
            # rather than stored against a paper of a different length. The
            # engine then falls back to the median of the years it does have.
            if canon_max[key].get(code) != cmax:
                skipped.append((key, year, code, cmax, canon_max[key].get(code)))
                continue
            # Each year is pro-rated against ITS OWN course total, which is what
            # makes a modified year usable at all: in 2023 Higher Biology Paper 2
            # was 95 marks of a 120-mark course, not 95 of 150.
            share = cmax / course_max
            a, b, c, d = (round(x * share) for x in (ba, bb, bc, bd))
            assert a > b > c > d > 0, "%s %s %s: %s not descending" % (
                sqa_name, year, code, (a, b, c, d))
            assert a < cmax, "%s %s %s: A %d is not below max %d" % (
                sqa_name, year, code, a, cmax)
            values.append(
                "    ('%s', 'SQA', '%s', '%s', 'June', NULL, %d, %d, %d, %d, NULL, TRUE)"
                % (key, code, year, a, b, c, d))
            seeds.append('    ("%s", "SQA", "%s", "%s", "June", None, %d, %d, %d),'
                         % (key, code, year, a, b, c))
    print("%s: %d subjects" % (year, len(SUBJECTS)))

if skipped:
    print("skipped (component absent or a different size that year):")
    for key, year, code, got, want in skipped:
        print("   %-16s %s %-16s max %s, catalogue says %s" % (key, year, code, got, want))

entries = []
for sqa_name, (key, label, colour) in sorted(SUBJECTS.items(), key=lambda kv: kv[1][0]):
    comps = shape[key]
    paper_lines = "\n".join(
        '                {"code": "%s", "name": "%s", "max_marks": %d%s},'
        % (code, lab, mx, "" if kind == "exam" else ', "assessment": "%s"' % kind)
        for code, lab, kind, mx in comps)
    topic_lines = "\n".join(
        '                "%s": [\n%s],' % (code, fmt(TOPICS[key][code]))
        for code, _l, _k, _m in comps)
    entry = '        "%s": {\n' % key
    entry += '            "name": "%s",\n' % label
    entry += '            "color": "%s",\n' % colour
    entry += '            "level": "Advanced Higher",\n'
    entry += ('            # Graded A-D: no A* and no E. The component max marks are\n'
              '            # SQA\'s own; the boundaries are derived from the course\n'
              '            # cut-off, because SQA publishes none per component.\n')
    entry += '            "papers": [\n%s\n            ],\n' % paper_lines
    entry += ('            # 2022 and 2023 ran in a modified form with the project\n'
              '            # removed, which is a different set of components.\n')
    entry += '            "years": ["SPEC", "2022", "2023", "2024", "2025"],\n'
    entry += '            "topics": {\n%s\n            },\n' % topic_lines
    entry += '        },\n'
    entries.append(entry)

sql_col = (
    "-- 037_boundary_derived_flag.sql\n"
    "-- Marks a boundary row as derived rather than published.\n"
    "--\n"
    "-- Every row in this table until now came from an awarding body's own\n"
    "-- document. The SQA rows do not: SQA publishes grade boundaries at course\n"
    "-- level only, so a component's boundary is that component's share of the\n"
    "-- course cut-off. That is an estimate, and an estimate that is\n"
    "-- indistinguishable from published data is a trap for whoever reads this\n"
    "-- table next.\n"
    "--\n"
    "-- Idempotent.\n\n"
    "ALTER TABLE grade_boundaries\n"
    "    ADD COLUMN IF NOT EXISTS derived_from_course BOOLEAN NOT NULL DEFAULT FALSE;\n"
)

sql = (
    "-- 038_sqa_advanced_higher_boundaries.sql\n"
    "-- SQA Advanced Highers: Biology, Chemistry, Physics, Mathematics,\n"
    "-- Economics, Geography, French, German, Spanish.\n"
    "--\n"
    "-- Graded A-D. a_star is NULL because there is no A*, and e_boundary is\n"
    "-- NULL because there is no E - below D is No Award. prediction.py reads\n"
    "-- both absences and builds a ladder that runs from D up to A.\n"
    "--\n"
    "-- derived_from_course is TRUE on every row. SQA publishes boundaries for\n"
    "-- the whole course only, never per component, so each component's\n"
    "-- boundary is its share of the course cut-off:\n"
    "--\n"
    "--     component_boundary = round(course_boundary * component_max / course_max)\n"
    "--\n"
    "-- The component max marks themselves are NOT derived - they are SQA's own,\n"
    "-- from the Assessment and Component Marks tables, and each course's\n"
    "-- components are checked to sum to the course maximum before anything is\n"
    "-- computed from them.\n"
    "--\n"
    "-- 2022-2025. Several of these courses ran in a modified form in 2022 and\n"
    "-- 2023 - coursework withdrawn, question papers resized - so a component\n"
    "-- carries a boundary for a year only when it was the same paper that\n"
    "-- year: same code, same max mark. Advanced Higher Physics' question\n"
    "-- paper was 155 marks then against 120 now, so it has no row for those\n"
    "-- years rather than one computed for a paper of a different length.\n"
    "-- Subjects that were not modified, English among them, carry all four.\n"
    "--\n"
    "-- Idempotent.\n\n"
    "-- Scoped to the Advanced Highers. When this was first written, SQA in\n"
    "-- this table meant only Advanced Highers, so deleting every SQA row was\n"
    "-- correct. It is not any more: migration 039 owns the Higher rows, and\n"
    "-- re-running this out of numeric order wiped them. A DELETE must never\n"
    "-- be wider than the INSERT that follows it.\n"
    "DELETE FROM grade_boundaries\n"
    "  WHERE board = 'SQA' AND subject LIKE '%% (AH)';\n\n"
    "INSERT INTO grade_boundaries\n"
    "    (subject, board, paper_code, year, series,\n"
    "     a_star, a_boundary, b_boundary, c_boundary, d_boundary, e_boundary,\n"
    "     derived_from_course)\n"
    "VALUES\n%s;\n"
) % (",\n".join(values))

with open(os.path.join(TELOS, "migrations", "037_boundary_derived_flag.sql"),
          "w", encoding="utf-8", newline="\n") as fh:
    fh.write(sql_col)
with open(os.path.join(TELOS, "migrations", "038_sqa_advanced_higher_boundaries.sql"),
          "w", encoding="utf-8", newline="\n") as fh:
    fh.write(sql)
with open(os.path.join(SP, "sqa_ah_seed.txt"), "w", encoding="utf-8") as fh:
    fh.write("\n".join(seeds))
with open(os.path.join(SP, "sqa_ah_entries.txt"), "w", encoding="utf-8") as fh:
    fh.write("".join(entries))
print("migrations 037/038 written with %d rows" % len(values))
